"""
Rapproche les 121 leçons du curriculum RAG des 101 compétences canoniques.

    python scripts/rapprocher_curriculum.py            # rapport à l'écran
    python scripts/rapprocher_curriculum.py --lot      # + classeur à faire valider

Pourquoi ce rapprochement
-------------------------
Le projet porte deux descriptions du programme qui ne se connaissent pas :

  · `data/knowledge/curriculum_*.yaml` — 121 leçons, riches en contenu
    pédagogique (savoir, savoir-faire, erreurs fréquentes) mais limitées au
    collège et indexées par un identifiant qui leur est propre (`4e_NUM_Ch4_L3`) ;
  · `02_Competences` du classeur — 101 compétences à code stable (`L.IDR`), avec
    prérequis chaînés et volume horaire, mais sans contenu rédigé.

Tant qu'ils restent disjoints, le module 7 (génération des fiches de remédiation)
n'a aucune matière à indexer : il connaîtra le problème à traiter — un couple
compétence × type d'erreur — sans pouvoir retrouver la leçon correspondante.

Ce que ce script fait, et ne fait pas
-------------------------------------
Il **propose**. Il ne décide pas. Le rapprochement est un jugement pédagogique :
« Repérage dans le plan » et « Coordonnées d'un point » peuvent désigner la même
compétence ou deux compétences distinctes selon le découpage retenu. Le script
classe donc ses propositions par niveau de confiance et sort un classeur à faire
valider — même discipline que pour les accents.

Deux signaux se sont révélés inutilisables, et c'est documenté ici pour éviter
qu'on les réessaie :

  · **les numéros de chapitre** — le curriculum écrit `Ch4`, le classeur `ch15`
    pour la même notion : les deux découpages ne coïncident pas ;
  · **les domaines** — le curriculum n'en connaît que deux (numérique,
    géométrique), le classeur huit. Le domaine ne sert qu'à écarter les
    rapprochements manifestement absurdes.

Reste le libellé de la leçon confronté à celui de la compétence, et la classe.
"""
from __future__ import annotations

import glob
import re
import sys
import unicodedata
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl
import yaml

_ROOT = Path(__file__).resolve().parent.parent
_KB = _ROOT / "data" / "knowledge"

# Une leçon de 3e peut relever d'une compétence introduite en 4e : les compétences
# sont introduites une fois puis reprises. On tolère donc un écart, sans l'ignorer.
_ORDRE_CLASSES = ["Primaire", "6eme", "5eme", "4eme", "3eme", "2ndeC", "1ereD"]
_CLASSE_CHUNK = {"6e": "6eme", "5e": "5eme", "4e": "4eme", "3e": "3eme"}

# Le curriculum ne connaît que deux domaines, le classeur huit. On s'en sert
# uniquement pour écarter l'absurde : une leçon géométrique ne devient pas une
# compétence de calcul littéral.
_DOMAINES_COMPATIBLES = {
    "numerique": {
        "Activites numeriques", "Calcul litteral et algebre", "Fonctions et applications",
        "Donnees, proportionnalite, statistique", "Suites et raisonnement",
        "Mesures et grandeurs",
    },
    "geometrique": {
        "Activites geometriques", "Mesures et grandeurs", "Trigonometrie",
    },
}

SEUIL_SUR = 0.52      # au-delà : proposition nette, à confirmer d'un coup d'œil
SEUIL_DOUTE = 0.34    # entre les deux : à trancher une par une

# Les deux sources ne nomment pas toujours la même notion de la même façon. Ces
# équivalences ont été relevées en confrontant les listes, pas devinées : sans
# elles, « Fonction affine » (curriculum) ne rejoignait pas « Applications
# lineaires et affines » (référentiel), faute d'un mot commun.
_SYNONYMES = {
    "fonction": "application",
    "fonctions": "application",
    "homothetie": "transformation",
    "translation": "transformation",
    "symetrie": "transformation",
    "frequence": "statistique",
    "frequences": "statistique",
    "moyenne": "statistique",
    "mediane": "statistique",
    "variance": "statistique",
    "effectif": "statistique",
    "effectifs": "statistique",
    "quotient": "division",
    "produit": "multiplication",
    "somme": "addition",
}

# Certaines compétences portent un libellé volontairement large (« Statistique du
# college ») et recouvrent plusieurs leçons du curriculum. La ressemblance
# textuelle ne peut pas les reconnaître : un libellé générique ne ressemble à
# aucune leçon précise. On les signale pour que le relecteur sache qu'un
# rapprochement large est attendu, plutôt que de conclure à une lacune.
_GENERIQUES = {"D.STAT1", "D.STAT2", "G.HOM", "F.REL", "D.TAB"}


def _fold(texte: str) -> str:
    sans_accents = "".join(
        c for c in unicodedata.normalize("NFD", str(texte or "")) if not unicodedata.combining(c)
    )
    return re.sub(r"[^a-z0-9]+", " ", sans_accents.lower()).strip()


def _mots(texte: str) -> set[str]:
    """Mots significatifs, avec les synonymes ajoutés à côté des originaux.

    Les mots très courts n'apportent rien et créent de faux rapprochements.
    """
    bruts = {m for m in _fold(texte).split() if len(m) > 3}
    return bruts | {_SYNONYMES[m] for m in bruts if m in _SYNONYMES}


def _similarite(chunk: dict, competence: dict) -> float:
    """Score composite, borné à 1.

    Le libellé de la leçon confronté à celui de la compétence pèse le plus : c'est
    le seul endroit où les deux sources parlent de la même chose dans les mêmes
    termes. Les mots-clés du chunk servent d'appoint.
    """
    lecon, libelle = _fold(chunk["lecon"]), _fold(competence["libelle"])
    score = SequenceMatcher(None, lecon, libelle).ratio()

    # Le libellé du classeur précise souvent la compétence après deux-points
    # (« Identites remarquables : developpement ») : on compare aussi au tronc.
    tronc = libelle.split(" : ")[0]
    if tronc != libelle:
        score = max(score, SequenceMatcher(None, lecon, tronc).ratio())

    mots_chunk = _mots(chunk["lecon"]) | set().union(
        *[_mots(m) for m in chunk.get("mots_cles", [])] or [set()]
    )
    mots_comp = _mots(competence["libelle"]) | _mots(competence["description"])
    if mots_chunk and mots_comp:
        recouvrement = len(mots_chunk & mots_comp) / len(mots_chunk)
        score = 0.7 * score + 0.3 * recouvrement

    # Classe : une correspondance exacte conforte, un écart de plus d'un cran
    # pénalise sans exclure.
    classe_chunk = _CLASSE_CHUNK.get(chunk["classe"])
    if classe_chunk and competence["niveau_intro"] in _ORDRE_CLASSES:
        try:
            ecart = abs(
                _ORDRE_CLASSES.index(classe_chunk)
                - _ORDRE_CLASSES.index(competence["niveau_intro"])
            )
        except ValueError:
            ecart = 0
        if ecart == 0:
            score += 0.08
        elif ecart > 1:
            score -= 0.06 * (ecart - 1)

    return max(0.0, min(1.0, score))


def _charger() -> tuple[list[dict], list[dict]]:
    chunks = [
        c
        for f in sorted(glob.glob(str(_KB / "curriculum_*.yaml")))
        for c in (yaml.safe_load(Path(f).read_text(encoding="utf-8")) or [])
    ]

    for candidat in (_ROOT.parent / "Referentiel_Socle_v0.xlsx", _ROOT / "Referentiel_Socle_v0.xlsx"):
        if candidat.exists():
            classeur = candidat
            break
    else:
        raise SystemExit("Referentiel_Socle_v0.xlsx introuvable.")

    wb = openpyxl.load_workbook(classeur, data_only=True, read_only=True)
    competences = [
        {
            "code": r[0], "domaine": r[1], "libelle": r[2], "description": r[3] or "",
            "niveau_intro": r[4], "chapitre_intro": r[5] or "",
        }
        for r in list(wb["02_Competences"].iter_rows(values_only=True))[1:]
        if r[0]
    ]
    wb.close()
    return chunks, competences


def rapprocher() -> dict:
    chunks, competences = _charger()
    surs, doutes, orphelins = [], [], []

    for chunk in chunks:
        compatibles = [
            c
            for c in competences
            if c["domaine"] in _DOMAINES_COMPATIBLES.get(chunk["domaine"], set())
        ] or competences

        classes = sorted(
            ((_similarite(chunk, c), c) for c in compatibles),
            key=lambda paire: paire[0],
            reverse=True,
        )
        meilleur, second = classes[0], (classes[1] if len(classes) > 1 else (0.0, None))

        entree = {
            "chunk": chunk,
            "propose": meilleur[1],
            "score": meilleur[0],
            "second": second[1],
            "score_second": second[0],
            # Un écart faible entre les deux meilleurs candidats signale une
            # ambiguïté réelle, même quand le score est bon.
            "ecart": meilleur[0] - second[0],
            "generique": meilleur[1]["code"] in _GENERIQUES,
        }

        if meilleur[0] >= SEUIL_SUR and entree["ecart"] >= 0.05:
            surs.append(entree)
        elif meilleur[0] >= SEUIL_DOUTE:
            doutes.append(entree)
        else:
            orphelins.append(entree)

    couverts = {e["propose"]["code"] for e in surs}
    sans_chunk = [c for c in competences if c["code"] not in couverts]

    return {
        "surs": surs, "doutes": doutes, "orphelins": orphelins,
        "competences_sans_chunk": sans_chunk,
        "total_chunks": len(chunks), "total_competences": len(competences),
    }


def _ecrire_lot(resultat: dict) -> Path:
    from datetime import date

    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    titre = Font(bold=True, size=11, color="FFFFFF")
    fond_titre = PatternFill("solid", fgColor="001E4A")
    fond_saisie = PatternFill("solid", fgColor="FFF6E5")
    haut = Alignment(vertical="top", wrap_text=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def feuille(nom: str, colonnes: list[tuple[str, int]], saisie_depuis: int):
        f = wb.create_sheet(nom)
        for i, (libelle, largeur) in enumerate(colonnes, start=1):
            cell = f.cell(row=1, column=i, value=libelle)
            cell.font, cell.fill, cell.alignment = titre, fond_titre, haut
            f.column_dimensions[get_column_letter(i)].width = largeur
        f.freeze_panes = "A2"
        return f, saisie_depuis

    def remplir(f, saisie_depuis, lignes):
        for index, valeurs in enumerate(lignes, start=2):
            for colonne, valeur in enumerate(valeurs, start=1):
                cell = f.cell(row=index, column=colonne, value=valeur)
                cell.alignment = haut
                if colonne >= saisie_depuis:
                    cell.fill = fond_saisie

    colonnes = [
        ("Leçon (curriculum)", 34), ("Classe", 8), ("Chapitre", 28),
        ("Identifiant", 20), ("Compétence proposée", 14), ("Libellé de la compétence", 40),
        ("Confiance", 11), ("Autre candidat", 14), ("Correct ?  ← O / N", 18),
        ("Si N : bon code  ← À REMPLIR", 26),
    ]

    f, s = feuille("01_A_confirmer", colonnes, 9)
    remplir(f, s, [
        (e["chunk"]["lecon"], e["chunk"]["classe"], e["chunk"]["chapitre"], e["chunk"]["id"],
         e["propose"]["code"], e["propose"]["libelle"], round(e["score"], 2),
         e["second"]["code"] if e["second"] else "", "", "")
        for e in sorted(resultat["surs"], key=lambda e: -e["score"])
    ])

    f, s = feuille("02_A_trancher", colonnes, 9)
    remplir(f, s, [
        (e["chunk"]["lecon"], e["chunk"]["classe"], e["chunk"]["chapitre"], e["chunk"]["id"],
         e["propose"]["code"], e["propose"]["libelle"], round(e["score"], 2),
         e["second"]["code"] if e["second"] else "", "", "")
        for e in sorted(resultat["doutes"], key=lambda e: -e["score"])
    ])

    # Une leçon sans proposition n'est pas forcément un échec de rapprochement :
    # le référentiel peut simplement ne pas couvrir la notion. La colonne le
    # demande explicitement, plutôt que de laisser le relecteur inventer un code.
    f, s = feuille("03_Sans_proposition", [
        ("Leçon (curriculum)", 34), ("Classe", 8), ("Chapitre", 28), ("Identifiant", 20),
        ("Savoir visé", 56), ("Compétence  ← code, ou « AUCUNE »", 30),
    ], 6)
    remplir(f, s, [
        (e["chunk"]["lecon"], e["chunk"]["classe"], e["chunk"]["chapitre"], e["chunk"]["id"],
         str(e["chunk"].get("savoir", ""))[:300], "")
        for e in resultat["orphelins"]
    ])

    # Notions enseignées par le curriculum et absentes du référentiel — vérifié
    # par recherche exhaustive dans les libellés ET descriptions des 101
    # compétences. Ce ne sont pas des échecs de rapprochement : aucune compétence
    # n'existe. Un élève échouant sur ces notions ne peut aujourd'hui pas être
    # diagnostiqué.
    f, s = feuille("05_Notions_non_couvertes", [
        ("Notion enseignée", 34), ("Classe", 10), ("Constat", 52),
        ("Décision  ← À REMPLIR", 40),
    ], 4)
    remplir(f, s, [
        ("Probabilité, dénombrement", "3e",
         "Aucune compétence au collège (D.DEN existe, mais en 1ère D)", ""),
        ("Angle inscrit, angle au centre", "3e",
         "Aucune compétence : les 6 compétences « angle » n'en traitent pas", ""),
        ("Similitude, figures semblables", "3e",
         "Aucune compétence ; G.THA (Thalès) est la plus proche sans recouvrir", ""),
        ("Variance, écart-type", "3e",
         "Aucune compétence ; D.STAT1 « Statistique du college » les couvre-t-elle ?", ""),
        ("Homothétie", "3e",
         "G.HOM existe mais est placée en 2nde C — décalage de niveau à trancher", ""),
    ])

    f, s = feuille("04_Competences_sans_lecon", [
        ("Code", 14), ("Libellé", 44), ("Domaine", 32), ("Niveau", 12), ("Remarque", 40),
    ], 6)
    remplir(f, s, [
        (c["code"], c["libelle"], c["domaine"], c["niveau_intro"],
         "Lycée : hors couverture du curriculum (6e-3e)"
         if c["niveau_intro"] in ("2ndeC", "1ereD") else "")
        for c in resultat["competences_sans_chunk"]
    ])

    sortie = _ROOT.parent / f"Lot_rapprochement_curriculum_{date.today().isoformat()}.xlsx"
    wb.save(sortie)
    return sortie


def main() -> int:
    resultat = rapprocher()

    print(f"{resultat['total_chunks']} leçons du curriculum · "
          f"{resultat['total_competences']} compétences canoniques\n")
    print(f"  À confirmer (proposition nette)   {len(resultat['surs']):4d}")
    print(f"  À trancher  (ambigu)              {len(resultat['doutes']):4d}")
    print(f"  Sans proposition                  {len(resultat['orphelins']):4d}")
    print(f"  Compétences sans leçon associée   {len(resultat['competences_sans_chunk']):4d}")

    lycee = sum(
        1 for c in resultat["competences_sans_chunk"]
        if c["niveau_intro"] in ("2ndeC", "1ereD")
    )
    if lycee:
        print(f"    dont lycée (hors couverture)    {lycee:4d}")

    print("\n  Aperçu des propositions les plus nettes :")
    for entree in sorted(resultat["surs"], key=lambda e: -e["score"])[:8]:
        print(f"    {entree['chunk']['id']:18s} {entree['chunk']['lecon'][:34]:34s} "
              f"→ {entree['propose']['code']:9s} ({entree['score']:.2f})")

    if "--lot" in sys.argv:
        sortie = _ecrire_lot(resultat)
        print(f"\n  Classeur à faire valider : {sortie}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
