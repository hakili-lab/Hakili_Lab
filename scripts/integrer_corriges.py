"""
Réintègre le lot complété par le relecteur pédagogique.

    python scripts/integrer_corriges.py --lot ../Lot_a_completer_2026-07-30.xlsx
    python scripts/integrer_corriges.py --lot ... --a-blanc    # rapport sans écriture

Écrit les corrigés dans `data/knowledge/corriges_socle.yaml`, **un fichier à part
des barèmes**. C'est délibéré : `generer_baremes_socle.py` régénère les barèmes
depuis le classeur source et écraserait des corrigés qui y seraient stockés. En
les gardant séparés, ce travail — plusieurs heures d'un enseignant — survit à
n'importe quelle régénération.

Le générateur de barèmes fusionne ce fichier s'il existe.

Les autres feuilles du lot (accents, volumes horaires, questions de construction)
ne sont pas réinjectées automatiquement : elles corrigent le **classeur source**,
qui appartient au docteur. Ce script les rapporte pour qu'on sache ce qui a été
rempli, et la correction se fait dans `Referentiel_Socle_v0.xlsx`.
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import yaml

_ROOT = Path(__file__).resolve().parent.parent
_KB = _ROOT / "data" / "knowledge"
_SORTIE = _KB / "corriges_socle.yaml"

_NIVEAUX = {
    "6ème": "6eme", "5ème": "5eme", "4ème": "4eme", "3ème": "3eme",
    "2nde C": "2ndeC", "1ère D": "1ereD", "Terminale D": "tleD",
}


def _texte(valeur) -> str:
    return "" if valeur is None else str(valeur).strip()


def _lire_corriges(wb) -> tuple[dict, list[str]]:
    if "01_Corriges" not in wb.sheetnames:
        return {}, ["Feuille 01_Corriges absente du lot."]

    corriges: dict[str, dict[str, dict]] = {}
    alertes: list[str] = []
    lignes = list(wb["01_Corriges"].iter_rows(values_only=True))[1:]

    for numero, ligne in enumerate(lignes, start=2):
        niveau_libelle, code = _texte(ligne[0]), _texte(ligne[1])
        if not code:
            continue
        niveau = _NIVEAUX.get(niveau_libelle)
        if niveau is None:
            alertes.append(f"Ligne {numero} : niveau « {niveau_libelle} » non reconnu.")
            continue

        reponse, demarche = _texte(ligne[7]), _texte(ligne[8])
        if not reponse and not demarche:
            continue  # non rempli : normal, on n'écrase rien
        if not reponse and demarche:
            alertes.append(
                f"Ligne {numero} ({niveau_libelle}/{code}) : démarche saisie sans "
                "réponse attendue — c'est la réponse qui sert à corriger."
            )
        corriges.setdefault(niveau, {})[code] = {
            "reponse_attendue": reponse,
            "solution": demarche,
        }

    return corriges, alertes


def _rapporter_autres(wb) -> None:
    """Les autres feuilles corrigent le classeur source, pas ce dépôt."""
    controles = [
        ("02_Accents", 6, "accents validés"),
        ("03_Accents_a_arbitrer", 5, "accents arbitrés"),
        ("04_Volumes_lycee", 6, "volumes horaires renseignés"),
        ("05_Questions_construction", 5, "critères de tracé renseignés"),
    ]
    for feuille, colonne, libelle in controles:
        if feuille not in wb.sheetnames:
            continue
        remplies = sum(
            1
            for ligne in list(wb[feuille].iter_rows(values_only=True))[1:]
            if len(ligne) >= colonne and _texte(ligne[colonne - 1])
        )
        total = max(0, wb[feuille].max_row - 1)
        if remplies:
            print(f"  {libelle:34s} {remplies}/{total}  → à reporter dans le classeur source")
        else:
            print(f"  {libelle:34s} 0/{total}")


def main() -> int:
    import argparse

    analyseur = argparse.ArgumentParser()
    analyseur.add_argument("--lot", required=True, help="Chemin du classeur complété")
    analyseur.add_argument("--a-blanc", action="store_true", help="Rapport sans écriture")
    options = analyseur.parse_args()

    chemin = Path(options.lot).expanduser().resolve()
    if not chemin.exists():
        print(f"Classeur introuvable : {chemin}")
        return 1

    wb = openpyxl.load_workbook(chemin, data_only=True, read_only=True)
    try:
        corriges, alertes = _lire_corriges(wb)
        print(f"Lot : {chemin}\n")
        total = sum(len(v) for v in corriges.values())
        for niveau in sorted(corriges):
            print(f"  {niveau:8s} {len(corriges[niveau]):3d} corrigé(s)")
        print(f"  {'TOTAL':8s} {total:3d}")
        print()
        _rapporter_autres(wb)
    finally:
        wb.close()

    if alertes:
        print()
        for alerte in alertes:
            print(f"  [ALERTE] {alerte}")

    if not total:
        print("\nAucun corrigé saisi — rien à écrire.")
        return 0

    if options.a_blanc:
        print("\nMarche à blanc : aucune écriture.")
        return 0

    # Fusion avec l'existant : un lot partiel ne doit pas effacer un travail
    # antérieur. Une valeur nouvellement saisie remplace l'ancienne.
    existant = {}
    if _SORTIE.exists():
        existant = yaml.safe_load(_SORTIE.read_text(encoding="utf-8")) or {}
        existant = existant.get("corriges", {})

    for niveau, questions in corriges.items():
        existant.setdefault(niveau, {}).update(questions)

    _SORTIE.write_text(
        "# Corrigés saisis à la main, hors du classeur source.\n"
        "# Fichier SÉPARÉ des barèmes, délibérément : generer_baremes_socle.py\n"
        "# régénère les barèmes depuis Referentiel_Socle_v0.xlsx et écraserait des\n"
        "# corrigés qui y seraient stockés. Il fusionne ce fichier à la génération.\n"
        "# Produit par scripts/integrer_corriges.py — modifiable à la main si besoin.\n\n"
        + yaml.safe_dump(
            {"corriges": existant}, allow_unicode=True, sort_keys=True, width=100
        ),
        encoding="utf-8",
    )

    conserve = sum(len(v) for v in existant.values())
    print(f"\nÉcrit : {_SORTIE.relative_to(_ROOT)} — {conserve} corrigé(s) au total.")
    print("Relancer ensuite : python scripts/generer_baremes_socle.py")
    return 0


if __name__ == "__main__":
    sys.exit(main())
