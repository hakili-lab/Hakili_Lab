"""
Produit le classeur de saisie à remettre au relecteur pédagogique.

    python scripts/generer_lot_a_completer.py

Rassemble en un seul fichier les trois manques qui bloquent le chantier et qui
demandent tous un enseignant de mathématiques — le même passage que la validation
du référentiel, déjà exigée par `00_Notice` (« À VALIDER ») :

  1. les 209 corrigés absents du référentiel (arbitrage B, chemin critique) ;
  2. les 150 mots à réaccentuer, plus 16 à arbitrer ;
  3. les 27 compétences de lycée sans volume horaire (arbitrage E).

Le classeur est **pré-rempli** : chaque ligne porte déjà le code de question, son
intitulé, sa compétence, son format et son barème. Il ne reste qu'à écrire la
réponse attendue — du remplissage, pas de la recherche.

Une fois rempli, `scripts/integrer_corriges.py` réinjecte le travail dans le
projet, et il devient insensible aux régénérations de barèmes.
"""
from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

import openpyxl
import yaml
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_ROOT = Path(__file__).resolve().parent.parent
_KB = _ROOT / "data" / "knowledge"

_TITRE = Font(bold=True, size=11, color="FFFFFF")
_FOND_TITRE = PatternFill("solid", fgColor="001E4A")
_FOND_SAISIE = PatternFill("solid", fgColor="FFF6E5")
_HAUT = Alignment(vertical="top", wrap_text=True)

_LIBELLES_NIVEAU = {
    "6eme": "6ème", "5eme": "5ème", "4eme": "4ème", "3eme": "3ème",
    "2ndeC": "2nde C", "1ereD": "1ère D", "tleD": "Terminale D",
}


def _entetes(feuille, colonnes: list[tuple[str, int]], *, saisie_depuis: int | None = None) -> None:
    for index, (libelle, largeur) in enumerate(colonnes, start=1):
        cellule = feuille.cell(row=1, column=index, value=libelle)
        cellule.font = _TITRE
        cellule.fill = _FOND_TITRE
        cellule.alignment = _HAUT
        feuille.column_dimensions[get_column_letter(index)].width = largeur
    feuille.freeze_panes = "A2"
    if saisie_depuis:
        feuille._saisie_depuis = saisie_depuis  # noqa: SLF001


def _notice(wb, comptes: dict) -> None:
    feuille = wb.create_sheet("00_Notice")
    feuille.column_dimensions["A"].width = 110
    lignes = [
        ("LOT À COMPLÉTER — Référentiel", True),
        (f"Généré le {date.today().isoformat()} depuis Referentiel_Socle_v0.xlsx", False),
        ("", False),
        ("Ce fichier rassemble ce qui manque au référentiel et qui demande un regard", False),
        ("d'enseignant. Les colonnes sur fond crème sont à remplir ; les autres sont là", False),
        ("pour vous situer et ne doivent pas être modifiées.", False),
        ("", False),
        ("01_Corriges — LE PLUS IMPORTANT", True),
        (f"{comptes['corriges']} questions sans réponse attendue. Le classeur ne donne la bonne", False),
        ("réponse que pour les QCM ; pour tout le reste, l'outil n'a rien à quoi comparer", False),
        ("la copie de l'élève. Tant que cette feuille n'est pas remplie, la correction", False),
        ("automatique ne fonctionne que sur un quart des questions.", False),
        ("", False),
        ("Deux colonnes à remplir :", False),
        ("  · Réponse attendue — le résultat, court et sans phrase. Ex. : « 147 », « 4x²−12x+9 »", False),
        ("  · Démarche attendue — facultative pour les questions courtes, utile pour les", False),
        ("    exercices rédigés où la démarche compte autant que le résultat.", False),
        ("", False),
        ("Vous aurez besoin des sujets PDF à côté : la colonne « Objet » décrit la question", False),
        ("mais ne la reproduit pas intégralement.", False),
        ("", False),
        ("02_Accents", True),
        (f"{comptes['accents']} mots écrits sans accent alors que leur forme accentuée existe", False),
        ("ailleurs dans le même classeur — ce sont des oublis, pas un choix. Ces textes", False),
        ("deviennent des libellés de question à l'écran et dans les rapports lus par les", False),
        ("parents.", False),
        ("", False),
        (f"03_Accents_a_arbitrer — {comptes['ambigus']} mots", True),
        ("Ici la forme sans accent est AUSSI un mot français valide : seule votre intention", False),
        ("tranche. « calcule » ou « calculé » ? « élève » (l'enfant) ou « élevé » (au carré) ?", False),
        ("Une correction automatique écrirait un contresens — d'où cette feuille séparée.", False),
        ("", False),
        ("04_Volumes_lycee — FACULTATIF", True),
        (f"{comptes['volumes']} compétences de lycée sans volume horaire officiel. Une valeur de repli", False),
        ("de 4 h leur est appliquée en attendant (médiane du collège), ce qui permet déjà", False),
        ("de calculer les coûts et les paliers. Cette feuille sert uniquement si vous", False),
        ("disposez des vrais volumes : elle n'est plus bloquante.", False),
        ("", False),
        ("05_Questions_construction", True),
        (f"{comptes['construction']} questions attendent un tracé géométrique, pas du texte. Elles ne pourront", False),
        ("vraisemblablement pas être corrigées automatiquement. Dites-nous ce qui doit être", False),
        ("vérifié sur le tracé, pour qu'un correcteur humain ait un critère clair.", False),
        ("", False),
        ("Une fois rempli", True),
        ("Renvoyez le fichier tel quel, sans le renommer ni changer l'ordre des colonnes.", False),
    ]
    for index, (texte, gras) in enumerate(lignes, start=1):
        cellule = feuille.cell(row=index, column=1, value=texte)
        cellule.alignment = Alignment(vertical="top", wrap_text=False)
        if gras:
            cellule.font = Font(bold=True, size=12 if index == 1 else 11)


def _corriges(wb) -> int:
    feuille = wb.create_sheet("01_Corriges")
    _entetes(
        feuille,
        [
            ("Niveau", 12), ("Code", 8), ("Partie", 8), ("Format", 14),
            ("Barème /20", 11), ("Compétence", 14), ("Objet de la question", 62),
            ("Réponse attendue  ← À REMPLIR", 34),
            ("Démarche attendue  ← facultatif", 46),
        ],
    )

    ligne = 2
    for niveau in _LIBELLES_NIVEAU:
        chemin = _KB / f"bareme_socle_{niveau}.yaml"
        if not chemin.exists():
            continue
        donnees = yaml.safe_load(chemin.read_text(encoding="utf-8"))
        for question in donnees["questions"]:
            if question["reponse_attendue"]:
                continue  # les QCM ont déjà leur bonne réponse
            valeurs = [
                _LIBELLES_NIVEAU[niveau],
                question["code"],
                question["partie"],
                question["format"],
                round(float(question["bareme"]), 4),
                question["competence"],
                question["objet"],
                "",
                "",
            ]
            for colonne, valeur in enumerate(valeurs, start=1):
                cellule = feuille.cell(row=ligne, column=colonne, value=valeur)
                cellule.alignment = _HAUT
                if colonne >= 8:
                    cellule.fill = _FOND_SAISIE
            ligne += 1
    return ligne - 2


def _accents(wb) -> tuple[int, int]:
    sys.path.insert(0, str(_ROOT / "scripts"))
    import importlib.util

    spec = importlib.util.spec_from_file_location(
        "verifier_accents", _ROOT / "scripts" / "verifier_accents.py"
    )
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    resultat = module.analyser(module._trouver_classeur())

    feuille = wb.create_sheet("02_Accents")
    _entetes(
        feuille,
        [("Mot écrit", 22), ("Corriger en", 24), ("Occurrences", 12),
         ("Déjà correct ailleurs", 20), ("Onglets", 40), ("Validé ?  ← À REMPLIR", 20)],
    )
    for index, entree in enumerate(resultat["a_corriger"], start=2):
        for colonne, valeur in enumerate(
            [entree["mot"], entree["suggestion"], entree["occurrences_sans"],
             entree["occurrences_avec"], ", ".join(entree["onglets"]), ""],
            start=1,
        ):
            cellule = feuille.cell(row=index, column=colonne, value=valeur)
            cellule.alignment = _HAUT
            if colonne == 6:
                cellule.fill = _FOND_SAISIE

    ambigus = wb.create_sheet("03_Accents_a_arbitrer")
    _entetes(
        ambigus,
        [("Mot écrit", 22), ("Une lecture possible", 24), ("Occurrences", 12),
         ("Onglets", 40), ("Forme correcte  ← À REMPLIR", 30)],
    )
    for index, entree in enumerate(resultat["a_arbitrer"], start=2):
        for colonne, valeur in enumerate(
            [entree["mot"], entree["suggestion"], entree["occurrences_sans"],
             ", ".join(entree["onglets"]), ""],
            start=1,
        ):
            cellule = ambigus.cell(row=index, column=colonne, value=valeur)
            cellule.alignment = _HAUT
            if colonne == 5:
                cellule.fill = _FOND_SAISIE

    return len(resultat["a_corriger"]), len(resultat["a_arbitrer"])


def _volumes_et_construction(wb, classeur: Path) -> tuple[int, int]:
    source = openpyxl.load_workbook(classeur, data_only=True, read_only=True)

    feuille = wb.create_sheet("04_Volumes_lycee")
    _entetes(
        feuille,
        [("Code", 14), ("Libellé", 46), ("Domaine", 32), ("Niveau", 10),
         ("Chapitre d'origine", 26), ("Volume horaire (h)  ← À REMPLIR", 30)],
    )
    n_volumes = 0
    for ligne in list(source["02_Competences"].iter_rows(values_only=True))[1:]:
        if not str(ligne[8]).lower().startswith("non disponible"):
            continue
        n_volumes += 1
        for colonne, valeur in enumerate(
            [ligne[0], ligne[2], ligne[1], ligne[4], ligne[5], ""], start=1
        ):
            cellule = feuille.cell(row=n_volumes + 1, column=colonne, value=valeur)
            cellule.alignment = _HAUT
            if colonne == 6:
                cellule.fill = _FOND_SAISIE

    construction = wb.create_sheet("05_Questions_construction")
    _entetes(
        construction,
        [("Niveau", 12), ("Code", 8), ("Compétence", 14), ("Objet de la question", 60),
         ("Ce qu'il faut vérifier sur le tracé  ← À REMPLIR", 52)],
    )
    n_constr = 0
    for ligne in list(source["04_Questions"].iter_rows(values_only=True))[1:]:
        if ligne[4] != "construction":
            continue
        n_constr += 1
        for colonne, valeur in enumerate(
            [_LIBELLES_NIVEAU.get(ligne[0], ligne[0]), ligne[1], ligne[6], ligne[8], ""],
            start=1,
        ):
            cellule = construction.cell(row=n_constr + 1, column=colonne, value=valeur)
            cellule.alignment = _HAUT
            if colonne == 5:
                cellule.fill = _FOND_SAISIE

    source.close()
    return n_volumes, n_constr


def main() -> int:
    import importlib.util

    spec = importlib.util.spec_from_file_location(
        "verifier_accents", _ROOT / "scripts" / "verifier_accents.py"
    )
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    classeur = module._trouver_classeur()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    n_corriges = _corriges(wb)
    n_accents, n_ambigus = _accents(wb)
    n_volumes, n_constr = _volumes_et_construction(wb, classeur)

    _notice(
        wb,
        {
            "corriges": n_corriges, "accents": n_accents, "ambigus": n_ambigus,
            "volumes": n_volumes, "construction": n_constr,
        },
    )
    # La notice en premier onglet, c'est la première chose qu'on ouvre.
    wb.move_sheet("00_Notice", offset=-len(wb.sheetnames) + 1)

    sortie = _ROOT.parent / f"Lot_a_completer_{date.today().isoformat()}.xlsx"
    wb.save(sortie)

    print(f"Classeur écrit : {sortie}\n")
    for libelle, n in [
        ("Corrigés à produire", n_corriges),
        ("Accents à corriger", n_accents),
        ("Accents à arbitrer", n_ambigus),
        ("Volumes horaires lycée", n_volumes),
        ("Questions de construction", n_constr),
    ]:
        print(f"  {libelle:28s} {n:4d}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
