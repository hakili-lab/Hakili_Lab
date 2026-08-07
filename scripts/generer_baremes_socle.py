"""
Génère les 7 barèmes des tests diagnostiques v2 depuis Referentiel_Socle_v0.xlsx.

Pourquoi ce script existe
-------------------------
Les 7 sujets PDF ne sont pas exploitables comme source de données : ils sont
générés par WeasyPrint et toutes les mathématiques y sont dessinées en vectoriel,
pas en texte. Une extraction PDF rend « Recopier et compléter avec le symbole
ou le symbole : et . » — la prose sans les formules.

Le classeur, lui, porte la même information en texte : `04_Questions` (code,
partie, barème, format, compétence canonique, intitulé) et `06_Distracteurs`
(les 4 options de chaque QCM, avec la bonne réponse). C'est donc lui la source.

Voir docs/harmonisation_donnees.md pour l'analyse complète.

Ce que le script produit
------------------------
data/knowledge/bareme_socle_<niveau>.yaml — un fichier par test, structure plate
(les 7 domaines N, L, G, D, F, M, S, T ne rentrent pas dans l'ancien découpage
`questions_numeriques` / `questions_geometriques`).

Idempotent : relancer régénère les mêmes fichiers à l'identique. À relancer après
toute mise à jour du classeur.

Ce que le script NE produit PAS
-------------------------------
Les corrigés des 209 questions non-QCM. Le classeur n'en contient aucun — c'est
l'arbitrage B de docs/harmonisation_donnees.md, non tranché. Les champs
`reponse_attendue` et `solution` sont émis vides pour ces questions, prêts à être
remplis sans changement de schéma. Les 71 QCM, eux, ont leur bonne réponse.

Usage :
    python scripts/generer_baremes_socle.py
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl
import yaml

_ROOT = Path(__file__).resolve().parent.parent
_KB_DIR = _ROOT / "data" / "knowledge"

# Le classeur vit à la racine du dossier de travail, au-dessus du dépôt.
_CLASSEUR_CANDIDATS = [
    _ROOT.parent / "Referentiel_Socle_v0.xlsx",
    _ROOT / "Referentiel_Socle_v0.xlsx",
]

# Barème : le classeur note sur 60 (30 questions de partie A à 1 pt,
# 10 exercices de partie B à 3 pts). Les sujets notent sur 20.
# Décision D-CEO-26 : le barème est stocké sur 20. Le rapport est exactement 3.
_DIVISEUR_BAREME = 3

# Niveaux du classeur → forme reconnue par src/core/classe_normalizer.py.
# Attention : "2ndeC", "1ereD", "TleD" ne sont PAS reconnus par normalize_classe —
# il faut les formes canoniques. La série (C/D) n'est pas distinguée par le
# normaliseur : "2nde C" devient "2nde".
_NIVEAUX: dict[str, dict[str, str]] = {
    "6eme":  {"classe": "6e",   "libelle": "6ème"},
    "5eme":  {"classe": "5e",   "libelle": "5ème"},
    "4eme":  {"classe": "4e",   "libelle": "4ème"},
    "3eme":  {"classe": "3e",   "libelle": "3ème"},
    "2ndeC": {"classe": "2nde", "libelle": "2nde C"},
    "1ereD": {"classe": "1ere", "libelle": "1ère D"},
    "tleD":  {"classe": "Tle",  "libelle": "Terminale D"},
}


def _trouver_classeur() -> Path:
    for p in _CLASSEUR_CANDIDATS:
        if p.exists():
            return p
    raise SystemExit(
        "Referentiel_Socle_v0.xlsx introuvable. Cherché dans :\n  "
        + "\n  ".join(str(p) for p in _CLASSEUR_CANDIDATS)
    )


def _tri_naturel(code: str) -> tuple[str, int]:
    """N1, N2, ..., N10 — et non N1, N10, N2."""
    m = re.match(r"([A-Z]+)(\d+)", code)
    return (m.group(1), int(m.group(2))) if m else (code, 0)


def _lire_classeur(chemin: Path) -> tuple[list, list, dict]:
    wb = openpyxl.load_workbook(chemin, data_only=True, read_only=True)
    questions = [r for r in wb["04_Questions"].iter_rows(values_only=True)][1:]
    distracteurs = [r for r in wb["06_Distracteurs"].iter_rows(values_only=True)][1:]
    competences = {
        r[0]: {"domaine": r[1], "libelle": r[2]}
        for r in list(wb["02_Competences"].iter_rows(values_only=True))[1:]
    }
    wb.close()
    return questions, distracteurs, competences


def _options_par_question(distracteurs: list) -> dict[tuple[str, str], list[dict]]:
    """(niveau, code_question) → options triées par lettre."""
    groupes: dict[tuple[str, str], list[dict]] = {}
    for r in distracteurs:
        niveau, code, lettre, texte, bonne, type_err, explication = r[:7]
        opt = {
            "lettre": str(lettre),
            "texte": "" if texte is None else str(texte),
            "correcte": str(bonne).strip().upper() == "OUI",
        }
        # Une bonne réponse ne révèle aucune erreur — les champs restent absents.
        if not opt["correcte"]:
            opt["type_erreur"] = None if type_err is None else str(type_err)
            opt["erreur"] = "" if explication is None else str(explication)
        groupes.setdefault((niveau, str(code)), []).append(opt)
    for key in groupes:
        groupes[key].sort(key=lambda o: o["lettre"])
    return groupes


def _charger_corriges_manuels() -> dict:
    """Corrigés saisis à la main, hors classeur source.

    Le classeur ne fournit de bonne réponse que pour les QCM. Les 209 autres sont
    produites par un relecteur pédagogique et vivent dans un fichier séparé
    (`corriges_socle.yaml`), précisément pour que cette régénération ne les efface
    pas. Voir scripts/integrer_corriges.py.
    """
    chemin = _KB_DIR / "corriges_socle.yaml"
    if not chemin.exists():
        return {}
    donnees = yaml.safe_load(chemin.read_text(encoding="utf-8")) or {}
    return donnees.get("corriges", {})


def _construire_test(
    niveau: str, questions: list, options: dict, competences: dict,
    corriges_manuels: dict | None = None,
) -> tuple[dict, dict]:
    meta_niveau = _NIVEAUX[niveau]
    manuels = (corriges_manuels or {}).get(niveau, {})
    lignes = sorted(
        (r for r in questions if r[0] == niveau), key=lambda r: _tri_naturel(str(r[1]))
    )

    items: list[dict] = []
    stats = {"qcm": 0, "sans_corrige": 0, "bareme_classeur": 0.0, "bareme_20": 0.0}

    for r in lignes:
        code = str(r[1])
        partie, bareme_classeur, fmt = str(r[2]), float(r[3]), str(r[4])
        comp_principale = str(r[6])
        comp_secondaire = None if r[7] is None else str(r[7])
        objet = "" if r[8] is None else str(r[8])

        bareme_20 = round(bareme_classeur / _DIVISEUR_BAREME, 6)
        comp = competences.get(comp_principale, {})

        item: dict = {
            "code": code,
            "partie": partie,
            "format": fmt,
            "bareme": bareme_20,
            "bareme_classeur": bareme_classeur,
            "competence": comp_principale,
            "competence_secondaire": comp_secondaire,
            "domaine": comp.get("domaine", ""),
            "objet": objet,
        }

        opts = options.get((niveau, code))
        if fmt == "qcm" and opts:
            correctes = [o["lettre"] for o in opts if o["correcte"]]
            # Une seule bonne réponse attendue — signalé plutôt que deviné.
            item["reponse_attendue"] = correctes[0] if len(correctes) == 1 else ""
            if len(correctes) != 1:
                print(
                    f"  [ALERTE] {niveau}/{code} : {len(correctes)} bonne(s) réponse(s) "
                    f"dans 06_Distracteurs — reponse_attendue laissée vide."
                )
            item["solution"] = ""
            item["options"] = opts
            stats["qcm"] += 1
        else:
            # Corrigé saisi à la main s'il existe ; sinon vide (arbitrage B).
            manuel = manuels.get(code, {})
            item["reponse_attendue"] = manuel.get("reponse_attendue", "")
            item["solution"] = manuel.get("solution", "")
            if item["reponse_attendue"]:
                stats["corriges_manuels"] = stats.get("corriges_manuels", 0) + 1
            else:
                stats["sans_corrige"] += 1

        stats["bareme_classeur"] += bareme_classeur
        stats["bareme_20"] += bareme_20
        items.append(item)

    meta = {
        "test_id": f"socle_{niveau}",
        "titre": f"Test diagnostique d'entrée en {meta_niveau['libelle']}",
        "niveau_test": niveau,
        "classe": meta_niveau["classe"],
        "total_questions": len(items),
        # Échelle /20 (décision D-CEO-26). La somme réelle des barèmes vaut
        # 19.99999… car une question de partie A vaut 1/3 de point : la note doit
        # toujours se calculer contre la somme réelle des max_score, jamais contre
        # ce total déclaré — sinon une copie parfaite ne vaut pas 20/20.
        "total_possible": 20,
        "somme_baremes_reelle": round(stats["bareme_20"], 6),
        "bareme_classeur_total": stats["bareme_classeur"],
        "source": "Referentiel_Socle_v0.xlsx — onglets 04_Questions, 06_Distracteurs, 02_Competences",
        "genere_par": "scripts/generer_baremes_socle.py",
        "note_corriges": (
            f"{stats['qcm']} QCM avec bonne réponse ; {stats['sans_corrige']} questions "
            "sans corrigé (absent du classeur — arbitrage B de docs/harmonisation_donnees.md). "
            "Les champs reponse_attendue et solution sont prêts à être remplis."
        ),
    }
    return {"meta": meta, "questions": items}, stats


def main() -> int:
    chemin = _trouver_classeur()
    print(f"Classeur : {chemin}\n")
    questions, distracteurs, competences = _lire_classeur(chemin)
    options = _options_par_question(distracteurs)

    inconnues = {
        str(r[6]) for r in questions if str(r[6]) not in competences
    }
    if inconnues:
        # Le module 0 a vérifié qu'il n'y en a aucune ; si ça change, on s'arrête.
        print(f"ERREUR : compétences absentes de 02_Competences : {sorted(inconnues)}")
        return 1

    corriges_manuels = _charger_corriges_manuels()
    if corriges_manuels:
        n = sum(len(v) for v in corriges_manuels.values())
        print(f"Corrigés saisis à la main repris : {n}\n")

    total_q = total_qcm = total_sans = total_manuels = 0
    for niveau in _NIVEAUX:
        data, stats = _construire_test(
            niveau, questions, options, competences, corriges_manuels
        )
        sortie = _KB_DIR / f"bareme_socle_{niveau}.yaml"
        sortie.write_text(
            "# Généré par scripts/generer_baremes_socle.py — ne pas éditer à la main.\n"
            "# Source : Referentiel_Socle_v0.xlsx. Relancer le script après toute\n"
            "# mise à jour du classeur. Exception : les champs reponse_attendue et\n"
            "# solution peuvent être complétés à la main (arbitrage B) — dans ce cas,\n"
            "# les sauvegarder avant de régénérer.\n\n"
            + yaml.safe_dump(
                data, allow_unicode=True, sort_keys=False, default_flow_style=False, width=100
            ),
            encoding="utf-8",
        )
        n = data["meta"]["total_questions"]
        print(
            f"  {sortie.name:28s} {n:3d} questions | {stats['qcm']:3d} QCM corrigés | "
            f"{stats['sans_corrige']:3d} sans corrigé | "
            f"barème {stats['bareme_classeur']:g}/60 → {stats['bareme_20']:.4f}/20"
        )
        total_q += n
        total_qcm += stats["qcm"]
        total_sans += stats["sans_corrige"]

    print(
        f"\nTotal : {total_q} questions · {total_qcm} QCM avec bonne réponse · "
        f"{total_sans} sans corrigé (arbitrage B)"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
