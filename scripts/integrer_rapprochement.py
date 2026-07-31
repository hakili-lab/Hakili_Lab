"""
Réintègre le rapprochement curriculum ↔ compétences, une fois validé.

    python scripts/integrer_rapprochement.py --lot ../Lot_rapprochement_curriculum_....xlsx
    python scripts/integrer_rapprochement.py --lot ... --a-blanc

Écrit un champ `code_competence` dans chaque leçon des `curriculum_*.yaml`.

Pourquoi écrire dans les fichiers du curriculum plutôt qu'ailleurs
------------------------------------------------------------------
Le rapprochement est une propriété de la leçon, pas une table de correspondance
séparée. En l'écrivant à côté de la leçon, `CurriculumRetriever` peut retourner
le contenu pédagogique indexé par code canonique sans jointure supplémentaire —
c'est exactement ce dont le module 7 a besoin pour générer une fiche de
remédiation à partir d'un couple compétence × type d'erreur.

Une leçon peut légitimement n'avoir aucune compétence : le curriculum enseigne des
notions que le référentiel ne couvre pas (probabilité, angle inscrit, similitude —
vérifié). Dans ce cas la valeur `AUCUNE` est acceptée et écrite telle quelle, pour
distinguer « pas encore rapproché » de « rapproché, et il n'y a rien ».
"""
from __future__ import annotations

import argparse
import glob
import sys
from pathlib import Path

import openpyxl
import yaml

_ROOT = Path(__file__).resolve().parent.parent
_KB = _ROOT / "data" / "knowledge"


def _texte(valeur) -> str:
    return "" if valeur is None else str(valeur).strip()


def _codes_valides() -> set[str]:
    for candidat in (_ROOT.parent / "Referentiel_Urie_v0.xlsx", _ROOT / "Referentiel_Urie_v0.xlsx"):
        if candidat.exists():
            wb = openpyxl.load_workbook(candidat, data_only=True, read_only=True)
            codes = {
                _texte(r[0])
                for r in list(wb["02_Competences"].iter_rows(values_only=True))[1:]
                if _texte(r[0])
            }
            wb.close()
            return codes
    raise SystemExit("Referentiel_Urie_v0.xlsx introuvable.")


def _lire_lot(chemin: Path, codes: set[str]) -> tuple[dict[str, str], list[str]]:
    """Retourne {identifiant_de_lecon: code_competence} et les alertes."""
    wb = openpyxl.load_workbook(chemin, data_only=True, read_only=True)
    associations: dict[str, str] = {}
    alertes: list[str] = []

    try:
        # 01 et 02 partagent la même disposition : proposition + validation.
        for feuille in ("01_A_confirmer", "02_A_trancher"):
            if feuille not in wb.sheetnames:
                continue
            for numero, ligne in enumerate(
                list(wb[feuille].iter_rows(values_only=True))[1:], start=2
            ):
                identifiant, propose = _texte(ligne[3]), _texte(ligne[4])
                verdict, correction = _texte(ligne[8]).upper(), _texte(ligne[9])
                if not identifiant:
                    continue

                if verdict.startswith("O"):
                    retenu = propose
                elif verdict.startswith("N"):
                    if not correction:
                        alertes.append(
                            f"{feuille} ligne {numero} ({identifiant}) : refusé sans "
                            "code de remplacement — leçon laissée non rapprochée."
                        )
                        continue
                    retenu = correction
                else:
                    continue  # non validé : on ne décide pas à sa place

                if retenu.upper() == "AUCUNE":
                    associations[identifiant] = "AUCUNE"
                elif retenu in codes:
                    associations[identifiant] = retenu
                else:
                    alertes.append(
                        f"{feuille} ligne {numero} ({identifiant}) : code « {retenu} » "
                        "absent du référentiel — ignoré."
                    )

        if "03_Sans_proposition" in wb.sheetnames:
            for numero, ligne in enumerate(
                list(wb["03_Sans_proposition"].iter_rows(values_only=True))[1:], start=2
            ):
                identifiant, saisi = _texte(ligne[3]), _texte(ligne[5])
                if not identifiant or not saisi:
                    continue
                if saisi.upper() == "AUCUNE":
                    associations[identifiant] = "AUCUNE"
                elif saisi in codes:
                    associations[identifiant] = saisi
                else:
                    alertes.append(
                        f"03_Sans_proposition ligne {numero} ({identifiant}) : code "
                        f"« {saisi} » absent du référentiel — ignoré."
                    )
    finally:
        wb.close()

    return associations, alertes


def main() -> int:
    analyseur = argparse.ArgumentParser()
    analyseur.add_argument("--lot", required=True)
    analyseur.add_argument("--a-blanc", action="store_true")
    options = analyseur.parse_args()

    chemin = Path(options.lot).expanduser().resolve()
    if not chemin.exists():
        print(f"Classeur introuvable : {chemin}")
        return 1

    codes = _codes_valides()
    associations, alertes = _lire_lot(chemin, codes)

    print(f"Lot : {chemin}\n")
    rapproches = sum(1 for v in associations.values() if v != "AUCUNE")
    aucune = sum(1 for v in associations.values() if v == "AUCUNE")
    print(f"  Leçons rapprochées à une compétence : {rapproches}")
    print(f"  Leçons déclarées sans compétence     : {aucune}")

    for alerte in alertes:
        print(f"  [ALERTE] {alerte}")

    if not associations:
        print("\nAucune validation saisie — rien à écrire.")
        return 0
    if options.a_blanc:
        print("\nMarche à blanc : aucune écriture.")
        return 0

    ecrits = 0
    for fichier in sorted(glob.glob(str(_KB / "curriculum_*.yaml"))):
        chemin_yaml = Path(fichier)
        lecons = yaml.safe_load(chemin_yaml.read_text(encoding="utf-8")) or []
        modifie = False
        for lecon in lecons:
            code = associations.get(lecon.get("id"))
            if code and lecon.get("code_competence") != code:
                lecon["code_competence"] = code
                modifie = True
                ecrits += 1
        if modifie:
            chemin_yaml.write_text(
                yaml.safe_dump(
                    lecons, allow_unicode=True, sort_keys=False, default_flow_style=False,
                    width=100,
                ),
                encoding="utf-8",
            )
            print(f"  {chemin_yaml.name} mis à jour")

    print(f"\n{ecrits} leçon(s) rapprochée(s) écrite(s) dans les curricula.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
